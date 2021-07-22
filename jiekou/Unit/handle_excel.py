import openpyxl
import os
import sys
base_path = os.getcwd()
sys.path.append(base_path)
# base_path = os.path.join(os.getcwd(), '../' + '/Config/'+'jiekou.xlsx')
# file_path = base_path + ".." + "/Config/jiekou.xlsx"
class HandExcel:

    def load_excel(self):
        open_excel = openpyxl.load_workbook(base_path+"/Config/jiekou.xlsx")
        return open_excel

    def get_sheet_data(self, index=0):
        sheet_name =self.load_excel().sheetnames
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self,row,cols):
        # 获取表格值
        data = self.get_sheet_data().cell(row,cols).value
        return data

    def get_rows(self):
        # 获取行数
        data = self.get_sheet_data().max_row
        return data

    def get_rows_value(self,row):
        # 获取某一行的内容
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def get_excel_data(self):
        # 获取excel里面所有的数据
        data_list = []
        for i in range(self.get_rows()-1):
            data_list.append(self.get_rows_value(i+2))
        return data_list

    def excel_write_data(self,row, cols, value):
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row,cols,value)
        wb.save(base_path+"/Config/jiekou.xlsx")

    def get_columns_value(self,key=None):
        '''
        获取某一列得数据
        '''
        columns_list = []
        if key==None:
            key = 'A'
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self,case_id):
        '''
        获取行号
        '''
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num+1
        return num

# def run_case():
#     rows = excel_data.get_rows()
#     for i in range(rows):
#         data = excel_data.get_rows_value(i+2)
#         print(data)


excel_data = HandExcel()
if __name__ == '__main__':
    excel_data.get_excel_data()